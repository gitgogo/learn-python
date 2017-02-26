#coding=utf-8
import re
s = 'aaa111aaa , bbb222 , 333ccc'
# 指定前后肯定断言
print re.findall( r'(?<=[a-z]{3})\d+(?=[a-z]+)', s) 

# 只指定后向肯定断言
print re.findall( r'\w+\d+(?=[a-z]+)', s) 

# 只指定前向肯定断言
print re.findall( r'(?<=[a-z]{3})\d+', s) 

# 普通匹配方法
print re.findall (r'[a-z]+(\d+)[a-z]+', s)
# 下面是一个错误的实例
try:
	matchResult = re.findall( r'(?<=[a-z]+)\d+(?=[a-z]+)', s) 
except Exception, e:
	print e
else:
	print matchResult

#xlsx
from openpyxl import Workbook
wb=Workbook()
ws1=wb.create_sheet(u'光荣之路1')
ws2=wb.create_sheet(u'光荣之路2')
ws3=wb.create_sheet(u'光荣之路3')
ws4=wb.create_sheet(u'光荣之路4')
for ws in [ws1,ws2,ws3,ws4]:
	ws['A1']='hello'
	ws.append(['world',23])

wb.save('f:\\test.xlsx')

#调整宽度
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Border,GradientFill
from openpyxl.styles import colors
from openpyxl import Workbook
from copy import copy

wb = Workbook()

sheet = wb.active

sheet['A1'] = 'Tall row'

sheet['B2'] = 'Wide column'

sheet.row_dimensions[1].height = 70

sheet.column_dimensions['B'].width = 20
wb.save(u"示例2.xlsx")
