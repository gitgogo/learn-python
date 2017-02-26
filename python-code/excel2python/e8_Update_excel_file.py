from openpyxl import Workbook,load_workbook
import datetime

wb=load_workbook(u'f:\\第一个文件.xlsx')

#获取第一个sheet
ws=wb.active
ws['A1']="gloryroad"
for i in range(4):
	print ws.rows[i][0].value

wb.save(u"f:\\第一个文件.xlsx")  #注意：将直接覆盖，不是更新。
