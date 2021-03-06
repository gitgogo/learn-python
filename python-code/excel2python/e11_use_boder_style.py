﻿from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook


wb = Workbook()
ws = wb.active
my_cell = ws['B2']
my_cell.value = "My Cell"
thin = Side(border_style="thin", color="BA55D3")
double = Side(border_style="double", color="BA55D3")
border = Border(top=double, left=thin, right=thin, bottom=double)
font = Font(b=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")
fill = PatternFill("solid", fgColor="E0FFFF")

my_cell.fill=fill
my_cell.border=border
my_cell.font=font
my_cell.alignment=al

wb.save(u"f:\\示例.xlsx")