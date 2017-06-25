#coding=utf-8
# from selenium import webdriver
# import time
from openpyxl import load_workbook

# wb=load_workbook(filename)
# ws=wb.active
# kw=ws.rows[1][0]
# ws.cell(row=2,column=ws.max_column+1,value='xxx')
# ws.rows[2][ws.max_column+1]='xxx'

# browser=webdriver.Chrome(executable_path='/Users/ralphliu/Document/webdriver/chromedriver')
# browser.get("https://www.baidu.com")
# with open('/Users/ralphliu/Document/test/data.txt') as f:
# 	for line in f:
# 		kw=line.split('||')[0].strip().decode('utf-8')
# 		ass=line.split('||')[1].strip().decode('utf-8')

# 		time.sleep(2)
# 		browser.find_element_by_id('kw').send_keys(kw)
# 		browser.find_element_by_id('su').click()
# 		time.sleep(3)
# 		assert ass in browser.page_source
# browser.quit()
# wb=load_workbook('test.xlsx')
# ws=wb.get_sheet_by_name('student')
# print ws.rows[2][4].value
# print ws.max_row,ws.max_column
# ws.append(['append'])
# ws.cell(row=2,column=4,value='cell')
# for cell in ws.rows:
# 	for i in range(ws.max_row):
# 		print cell[i].value
#coding=utf-8
# from xml.dom import minidom
# import codecs
# domTree=minidom.parse('interview.xml')
# root=domTree.documentElement
# print root.toxml()
# print root.getAttribute('name')
# # print root.hasAttribute('comment')
# position=root.getElementsByTagName('position')
# print position[0].getAttribute('comment')
# print position[0].childNodes[0].data
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

def send_email(body,attach):
	sender=""
	sender_pwd=""
	smtp_server=""
	receiver=""

	message=MIMEMultipart()
	message['From']=''
	message['To']=''
	message['Cc']=''
	message.attach(MIMEText(body,'plain','utf-8'))
	attach_file=open(attach,'rb')
	attach_part=MIMEApplication(attach_file.read(),Name=os.path.basename(attach))
	attach_part['Content-Disposition']="attachment: filename=%s"%os.path.basename(attach)
	message.attach(attach_part)

	try:
		smtp=smtplib.SMTP()
		smtp.connect(smtp_server)
		smtp.login(sender,sender_pwd)
		smtp.sendmail(sender,receiver,message.to_sting())
	except Exception as e:
		raise e